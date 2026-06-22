#!/usr/bin/env python3
"""
Auto-update the RBI NFS dashboard.

Steps:
  1. Download RBI's "Daily Settlement Data" workbook (handles the bot wall).
  2. Parse the NFS (through ATMs) Volume & Value columns from every monthly sheet.
  3. Splice the fresh data, a new version stamp, and a "published" label into index.html.

Designed to run unattended (e.g. from GitHub Actions). If the data has not
changed since the last build, index.html is left untouched and the script exits 0
so the scheduler simply commits nothing.

The dashboard (index.html) IS the template: this script only rewrites the three
constants  EMBEDDED_DATA / EMBEDDED_STAMP / EMBEDDED_PUBLISHED  inside it. All
styling, charts and logic are already baked into that file.
"""

import os, re, sys, json, time, datetime, io

RBI_FILE_URL  = "https://rbidocs.rbi.org.in/rdocs/content/docs/PSDDP04062020.xlsx"
RBI_REFERER   = "https://www.rbi.org.in/Scripts/FS_PaymentsData.aspx?fn=9"
HTML_PATH     = os.path.join(os.path.dirname(__file__), "index.html")
XLSX_PATH     = os.path.join(os.path.dirname(__file__), "_rbi_latest.xlsx")
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36")

MONTHS = {'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,'july':7,
          'august':8,'september':9,'october':10,'november':11,'december':12}
MABBR  = ['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']


# ---------------------------------------------------------------- fetch
def _looks_like_xlsx(b: bytes) -> bool:
    # .xlsx is a ZIP archive -> starts with 'PK'
    return bool(b) and b[:2] == b'PK'

def fetch_xlsx() -> bool:
    """Try a plain HTTP request first; fall back to a real headless browser
    (Playwright) which clears RBI's Imperva/Incapsula JavaScript challenge."""
    # Strategy 1 — requests with browser-like headers (works when not challenged)
    try:
        import requests
        headers = {"User-Agent": UA, "Referer": RBI_REFERER,
                   "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*"}
        r = requests.get(RBI_FILE_URL, headers=headers, timeout=90)
        if _looks_like_xlsx(r.content):
            open(XLSX_PATH, "wb").write(r.content)
            print("Fetched via requests (%d bytes)" % len(r.content)); return True
        print("requests path was challenged; falling back to headless browser…")
    except Exception as e:
        print("requests path failed:", e)

    # Strategy 2 — Playwright: visit the RBI page (sets the anti-bot cookie via JS),
    # then download the file inside the same browser context.
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(args=["--no-sandbox"])
            ctx = browser.new_context(user_agent=UA, accept_downloads=True)
            page = ctx.new_page()
            # 1) load the payments page so Incapsula issues a valid session cookie
            page.goto(RBI_REFERER, wait_until="networkidle", timeout=90000)
            page.wait_for_timeout(3000)
            # 2) fetch the file with the cookies now in the context
            resp = ctx.request.get(RBI_FILE_URL, headers={"Referer": RBI_REFERER}, timeout=90000)
            body = resp.body()
            browser.close()
            if _looks_like_xlsx(body):
                open(XLSX_PATH, "wb").write(body)
                print("Fetched via Playwright (%d bytes)" % len(body)); return True
            print("Playwright fetch did not return an xlsx (got %d bytes, head=%r)" % (len(body), body[:16]))
    except Exception as e:
        print("Playwright path failed:", e)
    return False


# ---------------------------------------------------------------- parse
def _num(x):
    if x is None: return None
    if isinstance(x, (int, float)): return float(x)
    s = str(x).strip()
    if s == "" or s.lower() == "h": return None
    try: return float(s.replace(",", ""))
    except Exception: return None

def _day(v, yr, mon):
    if isinstance(v, datetime.datetime): return v.day if v.month == mon and v.year == yr else None
    if isinstance(v, datetime.date):     return v.day if v.month == mon and v.year == yr else None
    s = str(v).strip()
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m: return int(m.group(3)) if int(m.group(2)) == mon and int(m.group(1)) == yr else None
    for fmt in ('%B %d, %Y', '%b %d, %Y', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            dt = datetime.datetime.strptime(s, fmt)
            return dt.day if dt.month == mon and dt.year == yr else None
        except Exception: pass
    return None

def parse_workbook(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    months, order = {}, []
    for sn in wb.sheetnames:
        name = sn.strip()
        mt = re.match(r'([A-Za-z]+)\s+(\d{4})$', name)
        if not mt: continue
        mon = MONTHS.get(mt.group(1).lower()); yr = int(mt.group(2))
        if not mon: continue
        rows = list(wb[sn].iter_rows(values_only=True))
        # locate the "NFS (through ATMs)" header dynamically (column drifts across years)
        nfs_col = hdr = None
        for ri in range(min(8, len(rows))):
            for j, c in enumerate(rows[ri] or []):
                if c and 'NFS' in str(c) and 'ATM' in str(c):
                    nfs_col, hdr = j, ri
        if nfs_col is None: continue
        vol_row = None
        for ri in range(hdr, min(hdr + 4, len(rows))):
            c = rows[ri][nfs_col] if nfs_col < len(rows[ri]) else None
            if c and str(c).strip().lower().startswith('vol'): vol_row = ri; break
        if vol_row is None: vol_row = hdr + 2
        days = []
        for ri in range(vol_row + 1, len(rows)):
            row = rows[ri] or []
            d0 = row[0] if row else None
            if d0 is not None and str(d0).strip().lower().startswith('note'): break
            dd = _day(d0, yr, mon)
            if dd is None: continue
            vol = _num(row[nfs_col])     if nfs_col   < len(row) else None
            val = _num(row[nfs_col + 1]) if nfs_col+1 < len(row) else None
            days.append({"d": dd, "vol": vol, "val": val})
        if days:
            key = f"{MABBR[mon]} {yr}"
            months[key] = {"year": yr, "month": mon, "days": days}
            order.append((yr, mon, key))
    order.sort()
    return months, [k for _, _, k in order]


# ---------------------------------------------------------------- splice
def _replace_between(src, start_marker, end_marker, new_inner):
    i = src.index(start_marker)
    j = src.index(end_marker, i + len(start_marker))
    return src[:i] + start_marker + new_inner + src[j:]

def splice(html, payload, stamp, published):
    # EMBEDDED_DATA spans one line, terminated by ';\nconst EMBEDDED_STAMP'
    html = _replace_between(html, "const EMBEDDED_DATA = ", ";\nconst EMBEDDED_STAMP",
                            json.dumps(payload, separators=(",", ":")))
    html = _replace_between(html, "const EMBEDDED_STAMP = ", ";", str(stamp))
    html = _replace_between(html, 'const EMBEDDED_PUBLISHED = "', '"', published)
    return html


# ---------------------------------------------------------------- main
def main():
    if not os.path.exists(HTML_PATH):
        print("ERROR: index.html (the dashboard template) is missing next to this script."); sys.exit(1)

    if not fetch_xlsx():
        print("ERROR: could not download the RBI file (bot protection / network). No changes made.")
        sys.exit(1)

    months, order = parse_workbook(XLSX_PATH)
    if not order:
        print("ERROR: no NFS data parsed from the workbook. No changes made."); sys.exit(1)

    lk = order[-1]
    maxday = max((d["d"] for d in months[lk]["days"] if d["vol"] is not None), default=0)
    published = f"{lk} (through day {maxday})"

    html = open(HTML_PATH, encoding="utf-8").read()

    # Has anything actually changed? Compare new data to what's already embedded.
    cur = re.search(r"const EMBEDDED_DATA = (\{.*?\});\nconst EMBEDDED_STAMP", html, re.S)
    new_months_json = json.dumps({"months": months, "monthOrder": order}, separators=(",", ":"))
    changed = True
    if cur:
        try:
            old = json.loads(cur.group(1))
            old_cmp = json.dumps({"months": old.get("months", {}), "monthOrder": old.get("monthOrder", [])},
                                 separators=(",", ":"))
            changed = (old_cmp != new_months_json)
        except Exception:
            changed = True
    if not changed:
        print("No change in RBI NFS data since last build — index.html left untouched.")
        try: os.remove(XLSX_PATH)
        except Exception: pass
        return

    payload = {
        "version": 1,
        "exportedAt": datetime.datetime.utcnow().isoformat() + "Z",
        "source": "RBI Daily Payment & Settlement — NFS (through ATMs) cash withdrawal",
        "units": {"volume": "lakh", "value": "INR Crore", "avgValue": "INR per transaction"},
        "months": months, "monthOrder": order,
    }
    stamp = int(time.time() * 1000)  # ms — strictly increasing, triggers viewers' auto-refresh
    html = splice(html, payload, stamp, published)
    open(HTML_PATH, "w", encoding="utf-8").write(html)
    print(f"Updated index.html — {len(order)} months, published: {published}, stamp: {stamp}")
    try: os.remove(XLSX_PATH)
    except Exception: pass


if __name__ == "__main__":
    main()
